import io, os, math, traceback
import streamlit as st
import pandas as pd
import openai
import requests
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import re

# ───────────────────────────────────────────────────────────────────────────────
# CONFIG
# ───────────────────────────────────────────────────────────────────────────────
if "OPENAI_API_KEY" not in st.secrets:
    st.error("OpenAI API key not found. Please add it to your Streamlit secrets.")
    st.stop()
openai.api_key = st.secrets["OPENAI_API_KEY"]

st.set_page_config(page_title="Fragrance Talkers - AI Corrector", layout="wide") # Updated page title for browser tab

# Initialize session state variables if they don't exist
if 'data_processed' not in st.session_state:
    st.session_state.data_processed = False
if 'excel_data_to_download' not in st.session_state:
    st.session_state.excel_data_to_download = None
if 'uploaded_file_name' not in st.session_state:
    st.session_state.uploaded_file_name = None
if 'out_name' not in st.session_state:
    st.session_state.out_name = "corrected_workbook.xlsx"
if 'error_occurred' not in st.session_state:
    st.session_state.error_occurred = False
if 'show_success_message' not in st.session_state:
    st.session_state.show_success_message = False

st.title("Fragrance Talkers – AI Brand and Product Name Correction & Data Transform") # Updated main heading
st.markdown(
    "Upload an Excel file containing fragrance data. This application will perform several data transformations, "
    "including SKU cleaning, type standardization, price formatting, and the generation of Offer, Status, Price, and Caveat columns. "
    "AI will be utilized to correct 'Brand' names and extract clean product names from 'Description' columns, preparing the data for talkers."
) # Updated description

MAX_ROWS_PER_SHEET = 500000
LARGE_SHEET_ROW_THRESHOLD = 10000
TRUNCATION_BUFFER = 20

# ───────────────────────────────────────────────────────────────────────────────
# HELPER FUNCTIONS
# ───────────────────────────────────────────────────────────────────────────────
def get_actual_column_name(df_columns, target_name):
    """Gets the actual column name from a list of columns, case-insensitive and space-insensitive."""
    # target_name is assumed to be clean (no leading/trailing spaces)
    for col in df_columns:
        if str(col).strip().lower() == target_name.lower(): # Added strip() here
            return col
    return None

def format_price_value(value):
    """Formats a price value: removes .00 for whole numbers, keeps other decimals (rounded to 2dp). Returns empty string for unformattable."""
    if pd.isna(value): # Catches actual NaN objects
        return ""
    
    value_str = str(value).strip()
    if not value_str: # Catches empty strings
        return ""
    
    # Explicitly return empty for "NA" or "N/A" strings before attempting numeric conversion
    if value_str.upper() in ["NA", "N/A"]:
        return ""

    match = re.search(r'(-?\s*[\d,]+(?:[.]\d+)?)', value_str)
    if match:
        num_str = match.group(1).replace(',', '').replace(' ', '')
        try:
            num_value = float(num_str)
            # Check if the number is effectively an integer (within a small tolerance for float precision)
            if abs(num_value - round(num_value)) < 1e-9 : 
                return str(int(round(num_value)))
            else:
                return f"{num_value:.2f}"
        except ValueError:
            return "" # Return empty if float conversion fails after regex match
    else:
        return "" # No number-like pattern found


# ───────────────────────────────────────────────────────────────────────────────
# BRAND LIST & ACRONYMS
# ───────────────────────────────────────────────────────────────────────────────
with st.spinner("Initializing application..."):
    MANUAL_BRANDS = sorted(list(set([
        "ABIB", "APLIN", "Abercrombie & Fitch", "Abib", "Acnecide", "Acqua di Parma", 
        "Adidas", "Advanced Nutrition Programme", "Aedes de Venustas", "Aeno", "Aeravida", "Aerin", 
        "Aēsop", "Aether", "Afnan", "Agent Provocateur", "AgooLus", "Aigner", "Air Val International", 
        "Ajmal", "Al Haramain", "Al Rehab", "Alaïa", "Alban Muller", "Alberto Balsam", "Alchimie Forever", 
        "Alexander McQueen", "Alexandre.J", "Alfa Italia", "Alfaparf", "Alfred Dunhill", "Alfred Sung", 
        "Algenist", "AllSaints", "Alpha-H", "Alpecin", "Altr London", "Alvarez Gomez", "Alyssa Ashley", 
        "Amouage", "American Crew", "Amika", "Amiot", "Amouroud", "Anastasia Beverly Hills", "Andreia Professional", 
        "Angel Schlesser", "Anna Sui", "Annayake", "Anne Möller", "Annick Goutal", "Anua", "Apothecary87", 
        "Aramis", "Ard Al Zaafaran", "Ardell", "Ariana Grande", "Armaf", "Armand Basi", "Armani", 
        "Aroma Works", "Aromatica", "Artdeco", "Artis", "Artiscent", "As I Am", "Aspect", "Aspen", 
        "Assert", "Aston Martin", "Astral", "Atelier Cologne", "Atelier des Ors", "Atelier PMP", 
        "Atkinsons", "Atrixo", "Aunt Jackie's", "Augustinus Bader", "Aura Cacia", "Australian Gold", 
        "Aussie", "Authentic Beauty Concept", "Aveda", "Aveeno", "Avène", "Avon", "Axiology", "Azzaro",
        "BPerfect Cosmetics", "BABE", "BARBER PRO", "BEAUTYPRO", "BEAUTYWORKS", "BIOSSANCE", "BLEACH LONDON", 
        "Babor", "Babyliss", "Bach", "Badedas", "Balance Activ", "Balance Me", "Balenciaga", "Balmain", 
        "Banana Boat", "Banana Republic", "Bare By Vogue", "bareMinerals", "Barry M", "Batiste", 
        "Bausch & Lomb", "Baxter of California", "Baylis & Harding", "Bazuka", "Beauté Pacifique", 
        "Beauty Bakerie", "Beauty Blender", "Beauty Formulas", "Beauty Pie", "BeautyPro", "BeautyStat", 
        "Beauty of Joseon", "Beckham", "Bed Head by TIGI", "Benefit", "Benefit Cosmetics", "Benetton", 
        "Bentley", "Berdoues", "Betty Barclay", "Beverly Hills Formula", "Beyu", "Bielenda", "Bill Blass", 
        "Billie Eilish", "Billion Dollar Brows", "Bio-Oil", "Bioderma", "Bioeffect", "Bioglan", "Biolage", 
        "Biotherm", "Biotulin", "Biretix", "Blinc", "Blistex", "Bloom and Blossom", "Blumarine", 
        "Bobbi Brown", "Boadicea The Victorious", "Body Fantasies", "Bodyform", "Bond No. 9", "Bondi Sands", 
        "BondiBoost", "Bon Parfumeur", "Bonjela", "Boost N Blend", "Borghese", "BosleyMD", "Botanics", 
        "Botanycl", "Bottega Veneta", "Bouclème", "Bourjois", "Bragg", "Braun", "Bread Beauty Supply", 
        "Briogeo", "Britney Spears", "Bronnley", "Brookland", "Brosse & Co", "Brow FX", "Brushworks", 
        "Brut", "Brylcreem", "Bulldog Skincare", "Bumble and bumble", "Burberry", "Burt's Bees", "BUSHBALM", 
        "Bvlgari", "By Terry", "Byredo", "Byphasse", "BYOMA", "By Rosie Jane", "byellie",
        "C.O.Bigelow", "CBDFX", "CCS", "CHANEL", "CHI", "CK One", "COCO & EVE", "COCO DE MER", "COOLA", "COSRX", 
        "Cacharel", "Cadiveu", "Calvin Klein", "Camille Rose", "Cantu", "Capasal", "Carbon Theory", 
        "Care", "Carita", "Carmex", "Carolina Herrera", "Caron", "Cartier", "Carven", "Catrice", 
        "Catwalk by TIGI", "Caudalie", "CeraVe", "Cerruti", "Cetaphil", "Cetraben", "Champo", 
        "Chantecaille", "Charles Worthington", "Charlie", "Charlotte Tilbury", "Charnos Hosiery", 
        "Chat D'Or", "Chella", "Childs Farm", "Chloé", "Chopard", "Christophe Robin", "Christina Aguilera", 
        "Cicabiafine", "Cien", "Clarins", "Clean", "Clean & Clear", "Clean Reserve", "Clearasil", 
        "Clearblue", "Clinique", "Clive Christian", "Clomana Beauty", "Cloud Nine", "Coach", "Colab", 
        "Colgate", "Collistar", "Collection", "Color Wow", "Color&Soin", "Colour Freedom", "Comme des Garçons", 
        "Compeed", "Comptoir Sud Pacifique", "Conair", "Condition Culture", "Conserving Beauty", 
        "Coola Suncare", "Copaci", "Corioliss", "Creed", "Curel", "Curlsmith", "Cutex", "Cuticura", "Cynos", "D:fi", 
        "DERMAFLASH", "DERMAdoctor", "DKNY", "Daimon Barber", "Daktarin", "Dame Products", "Danessa Myricks Beauty", 
        "Daniel Sandler", "Davidoff", "Davines", "David Beckham", "Dax", "Dear Dahlia", "Decleor", 
        "Deep Steep", "Delilah", "Delilah Chloe", "Denman", "Dennis Basso", "Dentek", "Dermalex", 
        "Dermalogica", "Dermatonics", "Dermedic", "Design.ME", "Dettol", "DevaCurl", "DHC", "Diesel", 
        "Difeel", "Difflam", "Dionis", "Dior", "Diptyque", "Disney", "Dolce & Gabbana", "Doll Beauty", 
        "Donna Karan", "Dose & Co", "Dot & Key", "Doublebase", "Dove", "Dr Botanicals", "Dr Brandt", 
        "Dr Hauschka", "Dr Irena Eris", "Dr Jart+", "Dr Organic", "Dr PawPaw", "Dr Sebagh", "Dr Squatch", 
        "Dr. Barbara Sturm", "Dr. Dennis Gross Skincare", "Dr. Lipp", "Dr. PAWPAW", "Drunk Elephant", 
        "Drybar", "Dunhill", "Dsquared2", "Dyson", "e.l.f. Cosmetics", "E45", "ECOOKING", "ECRU New York", 
        "ELEMIS", "ESPA", "Earth Harbor", "Earthly Body", "Ecooking", "EcoTools", "Edge", "Ed Hardy", 
        "Eight & Bob", "Elancyl", "Elchim", "Elegant Touch", "Elemis", "Elie Saab", "Elizabeth Arden", 
        "Elizabeth Taylor", "Ella Baché", "Ellips", "Ellis Brooklyn", "Elocon", "Emanuel Ungaro", "Embryolisse", 
        "Emjoi", "Emma Hardie", "Emporio Armani", "English Laundry", "Eos", "Epaderm", "Epitact", 
        "Erborian", "Ermenegildo Zegna", "Escada", "Escentric Molecules", "Essence", "Essie", "Estée Lauder", 
        "Etat Libre d'Orange", "Eternity", "Etro", "Eucerin", "Euthymol", "Eve Lom", "Eveline Cosmetics", 
        "EveryGreen", "Evo", "Evolve Beauty", "Ex Nihilo", "Eyeko", "Eylure",
        "FABCULE", "FUL", "Fabriq", "Faith in Nature", "Faithfull the Brand", "Fake Bake", "Fancii", 
        "Farmacy", "Farouk Systems", "Farsali", "Femfresh", "Fendi", "Fenjal", "Fenty Beauty", "Fenty Skin", 
        "Ferrari", "Fig.1", "Filorga", "Fine", "First Aid Beauty", "Fixodent", "Flawless", "Flexitol", 
        "Florence by Mills", "Floris", "Foamie", "Foreo", "Forest Essentials", "Formula 10.0.6", 
        "Fornasetti", "Franck Olivier", "Frank Body", "Frederic Malle", "Freederm", "Fresh", "Fruit Works", 
        "Fuente", "Fugazzi", "Fudge Professional", "Fudge Urban", "Fur",
        "GHOST", "GISOU", "GIVENCHY", "Gallinée", "Garnier", "Gatineau", "Geek & Gorgeous", "Gelish", 
        "Gengigel", "Geo. F. Trumper", "George Gina & Lucy", "Georgette Klinger", "Germolene", "Germoloids", 
        "Ghost", "ghd", "Gianfranco Ferre", "Gillette", "Gilly Hicks", "Giorgio Armani", "Giorgio Beverly Hills", 
        "Giovanni", "Girlactik", "Gisou", "Givenchy", "Glaze", "Glow Hub", "Glow Recipe", "Glossier", 
        "Gold Collagen", "Goldfaden MD", "Goldwell", "Good Molecules", "Good Patch", "Goodal", "Goop", 
        "Got2b", "Grande Cosmetics", "Grass & Co", "Green People", "Grown Alchemist", "Gruum", "Gucci", 
        "Guerlain", "Guess", "Guide Beauty", "Gun Ana", "Guy Laroche", "Gxve Beauty",
        "HASK", "HOLIKA HOLIKA", "Hair Recipe", "Hair Rituel by Sisley", "Hair Syrup", "HairGain", "HairBond", 
        "Hairburst", "Hairdo", "Hairmax", "Halston", "Hanz de Fuko", "Hawaiian Tropic", "Head & Shoulders", 
        "Head Jog", "Heath", "Heathcote & Ivory", "Hello Klean", "Hello Sunday", "Heliocare", "Hempz", 
        "Henna Plus", "Herbal Essences", "Hermès", "Hey Honey", "Histoires de Parfums", "Hollister", 
        "Holika Holika", "Honest Beauty", "Hot Tools", "Hourglass", "House of Lashes", "House of Oud", 
        "Huda Beauty", "Hugo Boss", "Human+Kind", "Hur", "Hush & Hush", "Huygens", "Hyaluronic", "Hydropeptide",
        "I Heart Revolution", "ICONIC London", "ILIA", "INIKA Organic", "INNOluxe", "ISDIN", "IT Cosmetics", 
        "Ikoo", "Illamasqua", "Iluminage", "Iman", "Indeed Labs", "Indie Lee", "Indola", "Indu", "Indulge", 
        "Inglot", "Initio Parfums Prives", "Inkdot", "Innisfree", "Institut Esthederm", "Invisibobble", 
        "Invity", "Irish Spring", "Isadora", "Isinis", "Isle of Paradise", "Issey Miyake",
        "J.F. Schwarzlose Berlin", "JLS", "JOICO", "JPG", "JVN Hair", "Jack Black", "Jacques Bogart", 
        "Jacques Fath", "Jaguar", "James Bond 007", "James Read", "Jane Iredale", "Jason", "Jason Wu", 
        "Jean Couturier", "Jean Desprez", "Jean Louis Scherrer", "Jean Patou", "Jean Paul Gaultier", 
        "Jennifer Aniston", "Jennifer Lopez", "Jenny Glow", "Jerome Russell", "Jessica McClintock", 
        "Jil Sander", "Jimmy Choo", "Jo Browne", "Jo Hansford", "Jo Malone London", "Joanna Vargas", 
        "John Frieda", "John Masters Organics", "John Varvatos", "Johnson's Baby", "Joico", "Jolen", 
        "Jones Road", "Joop!", "Jordan Samuel", "Jorgobé", "Jorum Studio", "Jouer Cosmetics", "Jovan", 
        "Juice Beauty", "Juicy Couture", "Juliette Has A Gun", "Jusbox", "Just For Men", "Juvena",
        "K18", "KMS", "KVD Beauty", "Kylie Cosmetics", "Kylie Jenner", "Kylie Skin", "Kai", "Kalms", 
        "Kantarō", "Kaplan MD", "Karl Lagerfeld", "Karmameju", "Kat Burki", "Kate Spade", "Kate Somerville", 
        "Kayali", "KayPro", "Kendamil", "Kenneth Cole", "Kensington", "Kent", "Kent Salon", "Kenzo", 
        "KeraCare", "KeraStraight", "Kerasilk", "Keratin Complex", "Kérastase", "Kevyn Aucoin", "Keys Soulcare", 
        "Kiehl's", "KIKO MILANO", "Kilian Paris", "Kim Kardashian", "King C. Gillette", "Kinship", "Kiss", 
        "Kitsch", "Kjaer Weis", "Klairs", "Klorane", "Koh Gen Do", "Kopari Beauty", "Kora Organics", "Korres", 
        "Kosas", "Kristin Ess", "Kumiko",
        "L'Artisan Parfumeur", "L'Erbolario", "L'OCCITANE", "L'Oréal", "L'Oréal Professionnel", "LA Girl", 
        "LA Splash", "LIHA Beauty", "La Biosthetique", "La Bouche Rouge", "La Collection Privée Christian Dior", 
        "La Perla", "La Prairie", "La Roche-Posay", "La Savonnerie Du Nouveau Monde", "Label.m", "Lacoste", 
        "Ladurée", "Lalique", "Lamisil", "Lancer", "Lancaster", "Lancôme", "Laneige", "Lanolips", "Lansinoh", 
        "Lanvin", "Lash FX", "Lash Perfect", "Lattafa", "Laura Biagiotti", "Laura Geller", "Laura Mercier", 
        "Le Couvent Maison de Parfum", "Le Galion", "Le Labo", "Le Mini Macaron", "Le Prunier", "Lee Stafford", 
        "Lemsip", "Lena Levi", "Lenti", "Leonor Greyl", "Les Bains Guerbois", "Les Eaux Primordiales", 
        "Lesquendieu", "Lethal Cosmetics", "Lqd", "Lierac", "Lil-Lets", "Lily Lolo", "Lime Crime", 
        "Linda Meredith", "Lipcote", "Listerine", "Little Butterfly London", "Live Tinted", "Living Proof", 
        "Liz Earle", "Loewe", "Lola James Harper", "Lolita Lempicka", "Lottie London", "Louis Vuitton", 
        "Love Beauty and Planet", "Love Boo", "Love Cosmetics", "Lovekins", "Lqd Skin Care", "Lucia Magnani", 
        "Lumene", "Lumi", "Luna Daily", "Lush", "Luxie", "Lynx", "lixirskin",
        "M.A.C", "MAC Cosmetics", "MCoBeauty", "MUA Makeup Academy", "MV Skintherapy", "Maaji", "Macadamia Natural Oil", 
        "Made by Mitchell", "Mad Hippie", "Madison Beer", "Madluvv", "Maëlys Cosmetics", "Maison Crivelli", 
        "Maison Francis Kurkdjian", "Maison Margiela", "Maison Matine", "Maison Tahité – Officine Creative Profumi", 
        "MakeUp Eraser", "Makeup Obsession", "Makeup Revolution", "Malibu", "Malin + Goetz", "MAM", "Mancera", 
        "Mane 'n Tail", "Manucurist", "Manuka Doctor", "Mar e Sol", "Marc Anthony", "Marc Jacobs", "Marc Weiss", 
        "Maria Nila", "Mario Badescu", "Marbert", "Marina de Bourbon", "Marvis", "Mary Cohr", "Masque BAR", 
        "Masqueology", "Matrix", "Maui Moisture", "Max Factor", "Maybelline", "Meder Beauty Science", "Medik8", 
        "Mediceuticals", "Megababe", "Memo Paris", "Mennace", "Mented Cosmetics", "Mercedes-Benz", "Merci Handy", 
        "Merit Beauty", "Michael Bublé", "Michael Kors", "Mielle Organics", "Mila Moursi", "Milani", 
        "Milk Makeup", "Miller Harris", "Mina", "MineTan", "Missguided", "Missoni", "Mitchum", "Miu Miu", 
        "Mixa", "Mizon", "Molton Brown", "Mon Guerlain", "Monday Haircare", "Monika Blunder", "Montana", 
        "Montblanc", "Montale", "Monu", "Moon", "Morphe", "Moroccanoil", "Moschino", "Mr Blanc Teeth", 
        "Mühle", "Mugler", "Murad", "Murdock London", "Mustela", "My Clarins", "Myvitamins", "Mylee",
        "N.C.P. Olfactives", "NAILSINC", "NARS", "NEST New York", "NIP+FAB", "NUXE", "NYX", "NYX Professional Makeup", 
        "Nailberry", "Nails Inc.", "Nak", "Naked Sundays", "Nanogen", "Nanokeratin", "Naomi Campbell", "Naomi Jon", 
        "Narciso Rodriguez", "Nascita", "Natasha Denona", "Natura Bissé", "Natura Siberica", "Natural Birthing Company", 
        "Natural Collection", "Nature Box", "Nature Spell", "Naturtint", "Nautica", "Neal & Wolf", "Neal's Yard Remedies", 
        "Necessaire", "Neom", "Neom Organics London", "NeoStrata", "Nesti Dante", "Neutrogena", "Neville", "NCLA Beauty", 
        "Nick Chavez", "Nico", "Nicolaï", "Nicole Scherzinger", "Nielsen", "Nina Ricci", "Nine Yards", "Nioxin", "Nivea", 
        "No Bleach London", "No7", "Noble Isle", "Nomad", "NoMo", "Noughty", "Novex", "Nudestix", "Nursem", "Nutrafol", 
        "Nutriganics", "Nutrisse",
        "OPI", "O&M", "OGX", "OUAI", "Oribe", "Olio E Osso", "Olay", "Olaplex", "Ole Henriksen", "Olew", "Olverum", 
        "Omorovicza", "Only Curls", "Oral-B", "Orchard", "Oribe Hair Care", "Origins", "Ormonde Jayne", "Orly", 
        "Oscar De La Renta", "Oskia", "Ouidad", "Oway", "PERFUME POD", # Added PERFUME POD
        "PUPA Milano", "Pacifica", "Paco Rabanne", "Pai Skincare", "Palmers", "Paloma Picasso", "Pantene", "Parfums de Marly", 
        "Paris Hilton", "Pat McGrath Labs", "Patchology", "Patrick Ta", "Pattern Beauty", "Paul Edmonds", "Paul Mitchell", 
        "Paul Smith", "Paula's Choice", "Peace Out Skincare", "Pearlie White", "Penhaligon's", "Pepe Jeans", "Percy & Reed", 
        "Percy Nobleman", "Perfectil", "Perricone MD", "Perris Monte Carlo", "Pestle & Mortar", "Peter Thomas Roth", 
        "Petite Amie Skincare", "Pharmaceris", "Philip B", "Philip Kingsley", "Philosophy", "Phlur", "Phytomer", "Phyto", 
        "Pierre Cardin", "Piggy Paint", "Pink Honey", "Pinrose", "Pixi", "Piz Buin", "Plant Apothecary", "Plantur", 
        "Playboy", "Police", "Polo Ralph Lauren", "Polished London", "Popband", "Porsche Design", "Positivf", "Prada", 
        "Prai Beauty", "Pravana", "Pregnacare", "Prem", "Pretty Athletic", "Primavera", "Priti NYC", "Private Boutique", 
        "Pro:Voke", "Proactiv", "Project Lip", "Proraso", "Prosper", "Prtz", "Pucci", "Pump Haircare", "Pure Fiji", 
        "Pure Romance", "Pureology", "Purito", "PYT Beauty",
        "Q+A", "Quai D'Orsay", "Queen Helene",
        "R+Co", "REN Clean Skincare", "ROSENTAL ORGANICS", "Radley", "Rahua", "Ralph Lauren", "RapidLash", "Rare Beauty", 
        "Raw Sugar", "Real Techniques", "Rebeluna", "Redken", "Regaine", "Regincos", "Relove by Revolution", "Rembrandt", 
        "Remescar", "René Furterer", "Revolution Pro", "Revolution Skincare", "RevitaLash", "Revlon", "Rhode Skin", 
        "Rihanna", "Rimmel", "Rituals", "RoC", "Roja Dove", "Roja Parfums", "Roots", "Rose & Caramel", "Rosie Jane", 
        "Rossano Ferretti Parma", "Route", "Routine", "Royal Apothic", "Rubis", "Ruby Hammer", "Rude Cosmetics",
        "SVR", "Sachajuan", "Sacred Nature", "Saie", "Saint Jane", "Sukin", "Sally Hansen", "Salt & Stone", "Salt Of The Earth", 
        "Salvatore Ferragamo", "Sam McKnight", "Sanctuary Spa", "Sand & Sky", "Sana Jardin", "Santa Maria Novella", 
        "Sarah Chapman", "Sarah Jessica Parker", "Sassoon", "Satin", "Savor Beauty", "Scaramouche & Fandango", 
        "Scentered", "Scholl", "Schwarzkopf", "Scientia", "Scott Barnes", "Sculpted by Aimee", "Sea Magik", "Sebamed", 
        "Seed Phytonutrients", "Selfless by Hyram", "Sensai", "Sense of Care", "Sensodyne", "Seoulista Beauty", 
        "Serge Lutens", "Sexy Hair", "Shaeri", "Shark Beauty", "Shay & Blue", "Shea Moisture", "Shhhowercap", 
        "Shikioriori", "Shiseido", "Shu Uemura", "Sigma Beauty", "Silke London", "Silkia", "Simple", "Simris", 
        "Sisley", "Sister & Co.", "Skin Authority", "Skin Chemists", "Skin Generics", "Skin Gym", "Skin Proud", 
        "Skin Research", "Skin Republic", "Skin Sapiens", "Skin Woof", "Skin Yoga", "SkinCeuticals", "Skinfix", 
        "Skyn Iceland", "Slip", "Smashbox", "Snooboos", "Soap & Glory", "So Eco", "Sofy", "Sol de Janeiro", 
        "Solgar", "St. Tropez", "Straand", "Studio London", "Summer Fridays", "Sunday Riley", "Sure",
        "Tangle Teezer", "Tan-Luxe", "Ted Baker", "The 7 Virtues", "The Body Shop", "The INKEY List", "The Ordinary",
        "Thierry Mugler", "Tom Ford", "Too Faced", "TRESemmé", "Trilogy", "TYPEBEA",
        "UKhair", "Umberto Giannini", "UNITE", "Urban Decay",
        "Valentino", "Valera", "Vichy", "Virtue", "Vita Liberata",
        "W7", "Weleda", "Wella Professionals", "WetBrush", "Womanizer", "Woodwick", "Woowoo",
        "Xerjoff",
        "YARDLEY", "YEPODA", "YSL", "Yardley", "Yes To", "Youth To The People", "Yves Saint Laurent", "Yves Rocher",
        "Zadig & Voltaire", "Zelens", "Zimmermann", "Zirh", "Zitsticka", "Zoella Beauty" 
    ])))

    ACRONYMS_TO_PRESERVE = sorted(list(set([
        "SPF", "UVA", "UVB", "PA+", "PA++", "PA+++", "PA++++", 
        "BB", "CC", "DD", "AM", "PM",
        "NYX", "DKNY", "CK", "MAC", 
        "EDP", "EDT", "EDC", 
        "N°5", "N°1", 
        "24K", "18K"
    ])))
    st.session_state.KNOWN_BRANDS = MANUAL_BRANDS
    st.session_state.ACRONYMS_TO_PRESERVE = ACRONYMS_TO_PRESERVE


PROMPT_TEMPLATE_BRAND = """
You are a brand name correction and formatting assistant. I will provide values from a **Brand** column of an Excel sheet. **Your tasks are to be performed in this specific order:**
1.  **Brand Identification & Correction**: Identify the primary brand name. Correct misspellings or abbreviations of this brand name to its proper form using the 'Known brands' list.
    * This includes expanding common brand abbreviations: "CK" to "Calvin Klein", "JPG" to "Jean Paul Gaultier", "SJP" to "Sarah Jessica Parker", "YSL" to "Yves Saint Laurent", "ZV" to "Zadig & Voltaire", and "GA" to "Giorgio Armani".
    * If an input brand name is already present in the 'Known brands' list (e.g., "Kylie Jenner", "Mar e Sol"), use the exact casing as it appears in the 'Known brands' list. Do not change a full known brand name to another full known brand name (e.g., do not change 'Kylie Jenner' to 'Kylie Cosmetics' if both are on the list).
2.  **If Additional Text is Present (Product Details, etc.)**: If the input cell contains more than just the primary brand name (e.g., it includes product names, sizes, fragrance types like initialisms PROFONDO EDT 50ml" alongside "GA"):
    * **Product Line Expansion (Crucial for mixed content in Brand column)**: Actively attempt to expand common product line initialisms or abbreviations (e.g., 'ADG', 'ADGH' for Giorgio Armani product lines such as 'Acqua di Gio' or 'Acqua di Gio Homme'; 'SWY' for 'Stronger With You') if they appear with an identified known brand. Use your knowledge of common product lines for the identified brand. The 'Known brands' list primarily contains main brand names, not all sub-lines.
    * **Fragrance Term Conversion**: Convert "Eau de Toilette" (and variations like "Eau du Toilette") to "EDT". Convert "Eau de Parfum" (and variations like "Eau du Parfum") to "EDP". "Parfum" by itself should remain "Parfum".
    * **Proper Case for Non-Brand Parts**: Apply 'Proper Case' (Title Case for most words) to the product name and other non-brand descriptive parts of the text.
    * **Acronym Preservation**: Ensure acronyms from the 'Acronyms to keep uppercase' list (like EDP, EDT, SPF) remain in ALL CAPS.
    * The goal is to return a fully corrected and well-formatted string, e.g., "GA ADGH PROFONDO EDT 50ml" should become "Giorgio Armani Acqua di Gio Homme Profondo EDT 50ml".
3.  **If Only Brand Name is Present**: If the input cell only contains the brand name (after potential correction/expansion from step 1), return just that corrected brand name.
4.  Leave all non-brand text or values that are clearly not brand names (e.g., product codes, general descriptions if they appear without a clear brand context) unchanged from their original form unless formatting rules (Proper Case, Acronyms) apply due to presence of a brand.

**Guidelines**
• Do **not** rephrase or add words unnecessarily, beyond the specified expansions and corrections.
• Preserve row order. Each input row in the CSV should correspond to an output row.
• Ensure the output CSV has the exact same number of rows as the input CSV.

**Known brands (use this list for corrections, paying attention to exact casing for brands like "Mar e Sol" and for primary brand identification):**
{brands}

**Acronyms to keep uppercase (ensure these appear in ALL CAPS if present, AFTER Proper Casing has been applied to the relevant parts of the text):**
{acronyms}

Return **only** the corrected data as a CSV with a single header `brand`. Do not include any other text, explanations, or markdown formatting like ```csv or ``` around the CSV output.

Input CSV to correct:
```csv
brand
{csv_data}
```
"""

PROMPT_TEMPLATE_DESCRIPTION = """
You are a product name extraction assistant. I will provide values from a **Description** column. Your goal is to extract ONLY the core product name, formatted for artwork.
**Perform the following steps in this exact order for EACH input row. Each step is mandatory and builds upon the previous one.**

1.  **Initial Brand & Product Line Correction/Expansion**:
    * **Identify Primary Brand**: First, identify any primary brand names (e.g., "Giorgio Armani", "Calvin Klein") within the full input description using the 'Known brands' list. Correct their spelling and casing to match the list (e.g., "Mar e Sol" should be "Mar e Sol"). Expand common primary brand abbreviations (e.g., "GA" to "Giorgio Armani", "YSL" to "Yves Saint Laurent", "CK" to "Calvin Klein").
    * **Expand Product Line Abbreviations (CRUCIAL AND MANDATORY)**: If a known primary brand is identified, **you MUST then expand any associated product line abbreviations or initialisms** to their full names. This is a critical step. For example:
        * If the input contains "GA ADGH PROFONDO", and "GA" is identified as "Giorgio Armani", you MUST expand "ADGH" to "Acqua di Gio Homme". The string at this stage becomes "Giorgio Armani Acqua di Gio Homme Profondo".
        * If "YSL L'HOMME" is found, and "YSL" is "Yves Saint Laurent", it becomes "Yves Saint Laurent L'Homme".
        * If "CK ONE SHOCK" is found, and "CK" is "Calvin Klein", it becomes "Calvin Klein One Shock".
    * Use your knowledge of common product lines for the identified brand. The objective is to form a "fully-corrected string" that includes the full primary brand and the full product line name. Let's call the output of this step the "Step 1 String".

2.  **Remove ONLY Primary Brand Name(s)**:
    * Take the "Step 1 String". From this string, you **MUST remove ONLY the full, corrected primary brand name(s)** that were identified and expanded in Step 1.
    * **Crucially, the expanded product line name (e.g., "Acqua di Gio Homme") MUST REMAIN.** Do NOT remove the product line.
    * **Example 1 (Handling "ADGH PROFONDO")**:
        * Input Description: "GA ADGH PROFONDO EDT 50ML"
        * Step 1 String: "Giorgio Armani Acqua di Gio Homme Profondo EDT 50ML"
        * After removing "Giorgio Armani", the string becomes: "Acqua di Gio Homme Profondo EDT 50ML". This is the "Step 2 String".
    * **Example 2**:
        * Input Description: "CK One Shock for Her EDT 100ml"
        * Step 1 String: "Calvin Klein One Shock for Her EDT 100ml"
        * After removing "Calvin Klein", the string becomes: "One Shock for Her EDT 100ml". This is the "Step 2 String".
    * If no primary brand was identified in Step 1, the "Step 1 String" passes through to become the "Step 2 String" unchanged.

3.  **Remove Fragrance & Product Type Terms (MANDATORY)**:
    * Take the "Step 2 String". From this string, you **MUST remove the following common fragrance type indicators if present**: "EDT", "Eau de Toilette", "Eau du Toilette", "EDP", "Eau de Parfum", "Eau du Parfum".
    * Also remove "Parfum" if it appears as a standalone type indicator (e.g., "XYZ Parfum 50ml" becomes "XYZ 50ml"). Be careful not to remove "Parfum" if it's clearly part of a product's actual name (e.g., "Le Parfum" should remain "Le Parfum" at this stage if it's the core name).
    * Additionally, remove terms like "Body Mist", "Body Spray", "Face Mist", "Setting Mist", "Hair Mist", "Mist", "Spray" if they appear to indicate the product *type or form* and are typically listed after the core product name, often near the size. Be cautious not to remove "Spray" or "Mist" if it's clearly an integral part of a unique product name itself.
    * Example: "Acqua di Gio Homme Profondo EDT 50ml" (the "Step 2 String") becomes "Acqua di Gio Homme Profondo 50ml". This is the "Step 3 String".

4.  **Remove Sizes (MANDATORY)**:
    * Take the "Step 3 String". From this string, you **MUST remove common product size indicators**. These can include patterns like:
        * "XXml", "XX ml", "XXML", "XX ML" (e.g., "50ml", "100 ml")
        * "X.Xfl oz", "X.X fl oz", "X.XFLOZ", "X.X FL OZ", "Xoz", "X oz" (e.g., "3.4fl oz", "1.7 FL. OZ.")
        * "XXg", "XX g", "XXGRAMS" (e.g., "15g", "200 g")
    * Example: "Acqua di Gio Homme Profondo 50ml" (the "Step 3 String") becomes "Acqua di Gio Homme Profondo". This is the "Step 4 String".

5.  **Clean and Format Remaining Text (Product Name) (MANDATORY)**:
    * Take the "Step 4 String". This text should be the core product name.
    * Trim any leading/trailing whitespace.
    * Apply 'Proper Case' (Title Case for most words) to this remaining product name.
    * If, after all removals, the string is empty or only whitespace, return an empty string.

6.  **Acronym Preservation in Final Product Name (MANDATORY)**:
    * If any acronyms from the 'Acronyms to keep uppercase' list are part of the *final, cleaned product name* from Step 5, ensure they remain in ALL CAPS (e.g., if a product name is "UV Plus Day Screen SPF50").

**Overall Processing Consistency**:
* You **MUST** apply ALL steps (1 through 6) diligently, sequentially, and accurately for EVERY row of input data. Do not skip steps or reduce the thoroughness of processing for later rows in a batch.
* The goal is to consistently transform each input description into a clean, formatted product name.

**Final Output Goal**: The result should be ONLY the cleaned and formatted product name, ready for use in artwork.

**Known brands (for Step 1 - primary brand identification):**
{brands}

**Acronyms to keep uppercase (for Step 6, applied to the final product name):**
{acronyms}

Return **only** the extracted and formatted product name as a CSV with a single header `description`. Do not include any other text, explanations, or markdown formatting like ```csv or ``` around the CSV output.

Input CSV to correct:
```csv
description
{csv_data}
```
"""

CHUNK_BRAND = 70
CHUNK_DESCRIPTION = 30
MODEL = "gpt-4o" # Changed from gpt-4o-mini

# UI elements defined once
status_text_global = st.empty()
prog_bar_global = st.empty()

file = st.file_uploader("Upload Excel Workbook (.xlsx)", type="xlsx", key="file_uploader_main")

# Reset state if file is removed
if not file and st.session_state.get('uploaded_file_name') is not None:
    st.session_state.data_processed = False
    st.session_state.excel_data_to_download = None
    st.session_state.uploaded_file_name = None
    st.session_state.out_name = "corrected_workbook.xlsx" # Reset out_name
    st.session_state.error_occurred = False
    st.session_state.show_success_message = False
    status_text_global.text("")
    prog_bar_global.empty()

if file is not None:
    # If a new file is uploaded or if the same file is re-processed (data_processed is False)
    if st.session_state.get('uploaded_file_name') != file.name or not st.session_state.data_processed:
        st.session_state.data_processed = False # Ensure it's false before processing starts
        st.session_state.excel_data_to_download = None
        st.session_state.uploaded_file_name = file.name
        st.session_state.error_occurred = False
        st.session_state.show_success_message = False # Reset success message flag

        base, ext = os.path.splitext(file.name)
        st.session_state.out_name = f"{base}_corrected{ext}"
        
        prog_bar = prog_bar_global.progress(0.0)
        status_text = status_text_global

        try:
            status_text.text("Reading Excel file...")
            workbook_original_read = pd.read_excel(file, sheet_name=None)
            
            with st.spinner("Pre-processing data (truncating large sheets, cleaning SKUs, standardizing types, formatting prices)..."):
                workbook = {}
                for sheet_name, df_orig in workbook_original_read.items():
                    df = df_orig.copy()
                    # Truncate large sheets logic
                    if df.shape[0] > LARGE_SHEET_ROW_THRESHOLD:
                        st.warning(f"Sheet '{sheet_name}' is large ({df.shape[0]} rows) and will be truncated for performance.")
                        df_dropped_all_nan = df.dropna(how='all').reset_index(drop=True)
                        if df_dropped_all_nan.empty:
                            df = df.iloc[:min(df.shape[0], TRUNCATION_BUFFER)].copy() if not df.empty else pd.DataFrame(columns=df.columns)
                        else:
                            max_idx_with_data = -1
                            for col in df_dropped_all_nan.columns:
                                try:
                                    series_str_stripped = df_dropped_all_nan[col].astype(str).str.strip()
                                    non_empty_series = series_str_stripped.replace('', pd.NA) # Treat empty strings as NA for last_valid_index
                                    last_valid = non_empty_series.last_valid_index()
                                    if last_valid is not None and last_valid > max_idx_with_data:
                                        max_idx_with_data = last_valid
                                except Exception as e_max_idx:
                                    st.warning(f"Sheet '{sheet_name}', col '{col}': Error finding last valid index: {e_max_idx}")
                            
                            if max_idx_with_data != -1:
                                new_row_count = max_idx_with_data + 1 + TRUNCATION_BUFFER
                                df = df_dropped_all_nan.iloc[:new_row_count].copy()
                            else: # If all rows were empty or only contained empty strings
                                df = df.iloc[:min(df.shape[0], TRUNCATION_BUFFER)].copy() if not df.empty else pd.DataFrame(columns=df.columns)
                    
                    # SKU code cleaning
                    sku_col_name = get_actual_column_name(df.columns, 'SKU code')
                    if sku_col_name:
                        try:
                            # Ensure column is string type for .str accessor
                            df[sku_col_name] = df[sku_col_name].astype(str)
                            # Create a boolean mask for rows where SKU is purely numeric
                            is_numeric_sku = df[sku_col_name].str.match(r'^\d+$', na=False) # na=False treats NaN as not matching
                            # Filter DataFrame to keep only rows with numeric SKUs
                            df = df[is_numeric_sku]
                        except Exception as e_sku:
                            st.warning(f"Sheet '{sheet_name}': Could not process 'SKU code' column: {e_sku}")

                    # Type standardization (EDT/EDP)
                    type_col_original_name = get_actual_column_name(df.columns, 'EDT/EDP')
                    type_col_final_name = 'Type' # Desired final name
                    size_col_name = get_actual_column_name(df.columns, 'Size')
                    if type_col_original_name:
                        df.rename(columns={type_col_original_name: type_col_final_name}, inplace=True)
                        if type_col_final_name in df.columns:
                            type_series = df[type_col_final_name].astype(str).str.strip().str.upper()
                            type_series = type_series.replace({'MIST': 'Body Mist', 'BODYMIST': 'Body Mist', 'BODY MIST': 'Body Mist'})
                            mask_each = (type_series == "EACH")
                            type_series[mask_each] = "" # Clear "EACH"
                            
                            if size_col_name and mask_each.any(): # If "EACH" was present and Size column exists
                                try:
                                    # For rows where type was "EACH", check if size is "1" or "1.0"
                                    size_values_to_check = df.loc[mask_each, size_col_name].astype(str).str.strip()
                                    condition_to_clear_size = (size_values_to_check == "1") | (size_values_to_check == "1.0")
                                    df.loc[mask_each & condition_to_clear_size, size_col_name] = "" # Clear these specific size values
                                except Exception as e_size:
                                     st.warning(f"Sheet '{sheet_name}': Error processing 'Size' for 'Each' type: {e_size}")
                            df[type_col_final_name] = type_series
                    
                    # Price formatting
                    price_cols_to_format = ['RRP', 'Now', 'Saving', 'WAS', 'UOM']
                    for col_base_name in price_cols_to_format:
                        actual_col = get_actual_column_name(df.columns, col_base_name)
                        if actual_col:
                            df[actual_col] = df[actual_col].apply(format_price_value)
                    workbook[sheet_name] = df.copy()

            jobs = []
            for sheet_name, df_processed in workbook.items():
                brand_col_name_actual = get_actual_column_name(df_processed.columns, "brand")
                desc_col_name_actual = get_actual_column_name(df_processed.columns, "description")
                if brand_col_name_actual:
                    try:
                        mask = df_processed[brand_col_name_actual].notna() & (df_processed[brand_col_name_actual].astype(str).str.strip() != '')
                        indices = df_processed.index[mask].tolist()
                        if indices: jobs.append({'sheet': sheet_name, 'column_name': brand_col_name_actual, 'column_type': 'brand', 'indices': indices, 'values': df_processed.loc[indices, brand_col_name_actual].astype(str).tolist()})
                    except KeyError: pass # Should not happen if get_actual_column_name worked
                if desc_col_name_actual:
                    try:
                        mask = df_processed[desc_col_name_actual].notna() & (df_processed[desc_col_name_actual].astype(str).str.strip() != '')
                        indices = df_processed.index[mask].tolist()
                        if indices: jobs.append({'sheet': sheet_name, 'column_name': desc_col_name_actual, 'column_type': 'description', 'indices': indices, 'values': df_processed.loc[indices, desc_col_name_actual].astype(str).tolist()})
                    except KeyError: pass

            status_text.text("Pre-processing complete.")
            if not jobs:
                st.error("No 'brand' or 'description' columns with processable data found after pre-processing.")
                st.session_state.error_occurred = True
            else:
                chunks_total = sum(math.ceil(len(job['values']) / (CHUNK_BRAND if job['column_type'] == 'brand' else CHUNK_DESCRIPTION)) for job in jobs)
                if chunks_total == 0: # Can happen if jobs list is not empty but all 'values' lists are empty
                    st.info("No actual data values to process with AI after filtering.")
                    # Proceed to final transformations and export if there's anything in workbook
                else:
                    status_text.text("Starting AI corrections...")
                    known_brands_str_sess = ", ".join(f'"{b}"' for b in st.session_state.KNOWN_BRANDS)
                    acronyms_str_sess = ", ".join(f'"{a}"' for a in st.session_state.ACRONYMS_TO_PRESERVE)
                    chunks_done = 0

                    for job_info in jobs:
                        sheet_name_job = job_info['sheet']
                        column_name_job = job_info['column_name']
                        column_type_job = job_info['column_type']
                        indices_to_update_job = job_info['indices']
                        values_to_process_job = job_info['values']
                        current_job_chunk_size = CHUNK_BRAND if column_type_job == 'brand' else CHUNK_DESCRIPTION
                        corrected_column_values_job = []
                        
                        if not values_to_process_job: # Skip if no values for this job
                            continue

                        total_chunks_for_job_val = math.ceil(len(values_to_process_job) / current_job_chunk_size)

                        for i_chunk, start_index_chunk in enumerate(range(0, len(values_to_process_job), current_job_chunk_size), start=1):
                            end_index_chunk = min(start_index_chunk + current_job_chunk_size, len(values_to_process_job))
                            current_chunk_for_llm_val = values_to_process_job[start_index_chunk:end_index_chunk]
                            status_text.text(f"AI Processing '{column_name_job}' in '{sheet_name_job}' (Chunk {i_chunk}/{total_chunks_for_job_val})")
                            csv_data_for_prompt_val = "\n".join([str(val).replace("\n", " ").replace("\r", " ") for val in current_chunk_for_llm_val])
                            
                            prompt_to_use = ""
                            if column_type_job == 'brand':
                                prompt_to_use = PROMPT_TEMPLATE_BRAND.format(brands=known_brands_str_sess, acronyms=acronyms_str_sess, csv_data=csv_data_for_prompt_val)
                            elif column_type_job == 'description':
                                prompt_to_use = PROMPT_TEMPLATE_DESCRIPTION.format(brands=known_brands_str_sess, acronyms=acronyms_str_sess, csv_data=csv_data_for_prompt_val)
                            else:
                                st.warning(f"Unknown column type '{column_type_job}' for AI processing in sheet '{sheet_name_job}', column '{column_name_job}'. Using original values for this chunk.")
                                corrected_column_values_job.extend(current_chunk_for_llm_val)
                                chunks_done += 1 # Increment as this chunk is "processed" by using original values
                                prog_bar.progress(chunks_done / chunks_total if chunks_total > 0 else 0)
                                continue # Skip API call for unknown type
                            
                            try:
                                response = openai.chat.completions.create(model=MODEL, messages=[{"role": "user", "content": prompt_to_use}], temperature=0)
                                llm_output_text = response.choices[0].message.content.strip()
                                if llm_output_text.lower().startswith("```csv"): llm_output_text = re.sub(r"^```csv\s*\n", "", llm_output_text, flags=re.IGNORECASE)
                                if llm_output_text.lower().startswith("```"): llm_output_text = re.sub(r"^```\s*\n", "", llm_output_text, flags=re.IGNORECASE) # Generic ```
                                if llm_output_text.endswith("```"): llm_output_text = llm_output_text[:-3].strip()
                                
                                lines = llm_output_text.splitlines()
                                # Expects first line to be header if present, or just data
                                parsed_list_from_llm = [line.strip() for line in (lines[1:] if lines and lines[0].strip().lower() == column_type_job else lines)]

                                if not parsed_list_from_llm and current_chunk_for_llm_val: # LLM returned empty for non-empty input
                                    st.warning(f"LLM returned empty response for chunk {i_chunk} of '{column_name_job}' in '{sheet_name_job}'. Using original values for this chunk.")
                                    parsed_list_from_llm = list(current_chunk_for_llm_val) # Use a copy

                                expected_len = len(current_chunk_for_llm_val)
                                if len(parsed_list_from_llm) != expected_len:
                                    st.warning(f"LLM output length mismatch for chunk {i_chunk} of '{column_name_job}' in '{sheet_name_job}'. Expected {expected_len}, got {len(parsed_list_from_llm)}. Adjusting/padding with original values.")
                                    if len(parsed_list_from_llm) > expected_len:
                                        parsed_list_from_llm = parsed_list_from_llm[:expected_len]
                                    else: # len(parsed_list_from_llm) < expected_len
                                        # Pad with original values from the current chunk
                                        padding = list(current_chunk_for_llm_val[len(parsed_list_from_llm):])
                                        parsed_list_from_llm.extend(padding)
                                corrected_column_values_job.extend(parsed_list_from_llm)
                            except Exception as e_ai:
                                st.error(f"Error during AI processing for chunk {i_chunk} of '{column_name_job}' in '{sheet_name_job}': {e_ai}. Using original values for this chunk.")
                                corrected_column_values_job.extend(current_chunk_for_llm_val) # Fallback to original
                            
                            chunks_done += 1
                            prog_bar.progress(chunks_done / chunks_total if chunks_total > 0 else 0)
                        
                        if len(indices_to_update_job) == len(corrected_column_values_job):
                            df_target = workbook[sheet_name_job]
                            for idx_val, corrected_val in zip(indices_to_update_job, corrected_column_values_job):
                                if idx_val not in df_target.index: # Should not happen if indices are from df_target
                                    st.error(f"Index {idx_val} for '{column_name_job}' in '{sheet_name_job}' is out of bounds. Skipping update for this value.")
                                    continue
                                try:
                                    df_target.at[idx_val, column_name_job] = corrected_val
                                except Exception as e_at:
                                    st.error(f"Error setting value at index {idx_val} for column '{column_name_job}' in sheet '{sheet_name_job}': {e_at}")
                        else:
                            st.error(f"Length mismatch between indices to update ({len(indices_to_update_job)}) and corrected values ({len(corrected_column_values_job)}) for '{column_name_job}' in '{sheet_name_job}'. This column was NOT updated with AI results.")

                status_text.text("AI corrections complete. Applying final data transformations...")
                with st.spinner("Applying final data transformations (Offer, Status, Price, Caveat)..."):
                    for sheet_name_final, df_final in workbook.items():
                        # Get actual column names once per sheet
                        offer_col = get_actual_column_name(df_final.columns, 'Offer')
                        saving_col = get_actual_column_name(df_final.columns, 'Saving')
                        now_col = get_actual_column_name(df_final.columns, 'Now')
                        rrp_col = get_actual_column_name(df_final.columns, 'RRP')
                        brand_col = get_actual_column_name(df_final.columns, 'Brand')
                        desc_col = get_actual_column_name(df_final.columns, 'Description')
                        type_col = get_actual_column_name(df_final.columns, 'Type') # Already renamed from EDT/EDP
                        size_col = get_actual_column_name(df_final.columns, 'Size')
                        uom_col = get_actual_column_name(df_final.columns, 'UOM')

                        # --- Offer Column Logic ---
                        if offer_col:
                            new_offer_values = []
                            for _, row in df_final.iterrows():
                                current_offer_input = row.get(offer_col)
                                current_offer_str = str(current_offer_input).strip() if pd.notna(current_offer_input) else ""
                                current_offer_upper = current_offer_str.upper()
                                
                                processed_offer = current_offer_str # Default to original

                                rrp_val_from_row = str(row.get(rrp_col, "")).strip() if rrp_col else ""
                                saving_val_formatted = str(row.get(saving_col, "")).strip() if saving_col else ""


                                if "SAVE VS RRP" in current_offer_upper or "SAVE V RRP" in current_offer_upper:
                                    if saving_val_formatted: 
                                        processed_offer = f"save £{saving_val_formatted} on RRP"
                                elif current_offer_str.lower().startswith("save ") and current_offer_str.endswith("%"):
                                    processed_offer = "save " + current_offer_str[5:] 
                                elif current_offer_upper == "SAVE":
                                    if saving_val_formatted:
                                        processed_offer = f"save £{saving_val_formatted}"
                                elif current_offer_str == "" or current_offer_upper in ["NA", "N/A"]: 
                                    if rrp_val_from_row: 
                                        processed_offer = f"RRP £{rrp_val_from_row}"
                                    else: 
                                        processed_offer = "" 
                                else: 
                                    is_already_save_on_rrp = processed_offer.lower().startswith("save £") and "on rrp" in processed_offer.lower()
                                    is_already_save_percent = processed_offer.lower().startswith("save ") and processed_offer.endswith("%")
                                    is_already_save_amount = processed_offer.lower().startswith("save £") and not ("on rrp" in processed_offer.lower())
                                    is_already_rrp_amount = processed_offer.upper().startswith("RRP £")

                                    if not (is_already_save_on_rrp or is_already_save_percent or is_already_save_amount or is_already_rrp_amount):
                                        processed_offer = current_offer_str.lower()
                                
                                new_offer_values.append(processed_offer)
                            df_final[offer_col] = new_offer_values
                        
                        # --- Status Column Logic ---
                        df_final['Status'] = "" 
                        if offer_col: 
                            for idx, row_status in df_final.iterrows(): 
                                current_offer_val = str(row_status.get(offer_col, '')) 
                                current_offer_lower = current_offer_val.lower()
                                
                                status_to_set = ""
                                if current_offer_lower.startswith("save £") and "on rrp" in current_offer_lower:
                                    status_to_set = "only"
                                elif current_offer_lower.startswith("save ") and current_offer_val.endswith("%"): 
                                    status_to_set = "now"
                                elif current_offer_lower.startswith("save £"):
                                    status_to_set = "now"
                                elif "only" in current_offer_lower: 
                                    if not (current_offer_lower.startswith("save £") and "on rrp" in current_offer_lower):
                                        status_to_set = "only"
                                elif current_offer_val.upper().startswith("RRP"): 
                                    status_to_set = "RRP"
                                df_final.loc[idx, 'Status'] = status_to_set

                        # --- Price Column Logic ---
                        df_final['Price'] = "" 
                        for idx, row_price in df_final.iterrows():
                            status = str(row_price.get('Status', ''))
                            price_val_str = ""
                            now_val_from_row = str(row_price.get(now_col, "")).strip() if now_col else ""
                            rrp_val_from_row_for_price = str(row_price.get(rrp_col, "")).strip() if rrp_col else ""

                            if status in ["now", "only"] and now_val_from_row:
                                price_val_str = now_val_from_row 
                            elif status == "RRP" and rrp_val_from_row_for_price:
                                price_val_str = rrp_val_from_row_for_price 
                            
                            df_final.loc[idx, 'Price'] = f"£{price_val_str}" if price_val_str else ""

                        # --- Caveat Column Logic ---
                        df_final['Caveat'] = "" 
                        for idx, row_caveat in df_final.iterrows():
                            b = str(row_caveat.get(brand_col, "")).strip() if brand_col else ""
                            d = str(row_caveat.get(desc_col, "")).strip() if desc_col else ""
                            t = str(row_caveat.get(type_col, "")).strip() if type_col else ""
                            s_val = str(row_caveat.get(size_col, "")).strip() if size_col else "" 
                            
                            uom_price_val = str(row_caveat.get(uom_col, "")).strip() if uom_col else "" 

                            parts = []
                            if b and d: parts.append(f"{b} {d}")
                            elif b: parts.append(b)
                            elif d: parts.append(d)
                            
                            type_size_parts = []
                            if t: type_size_parts.append(t)
                            if s_val: type_size_parts.append(s_val) 
                            if type_size_parts: parts.append(" ".join(type_size_parts))

                            if uom_price_val: parts.append(f"£{uom_price_val} per 100ml") 
                            
                            caveat_text = ", ".join(filter(None, parts))
                            if caveat_text: caveat_text += ". "
                            caveat_text += "Subject to availability. Selected lines only."
                            df_final.loc[idx, 'Caveat'] = caveat_text
                        
                        workbook[sheet_name_final] = df_final.copy() 

                status_text.text("Final transformations complete. Preparing for export...")
                export_data_buffer = io.BytesIO()
                with pd.ExcelWriter(export_data_buffer, engine="openpyxl") as writer_export:
                    for sheet_name_export, df_export in workbook.items():
                        if df_export.empty:
                            st.info(f"Sheet '{sheet_name_export}' is empty. Skipping export for this sheet.")
                            continue
                        if df_export.shape[0] > MAX_ROWS_PER_SHEET:
                            st.warning(f"Sheet '{sheet_name_export}' has {df_export.shape[0]} rows, exceeding max per sheet ({MAX_ROWS_PER_SHEET}). It will be split.")
                            for i_part, start_row_part in enumerate(range(0, df_export.shape[0], MAX_ROWS_PER_SHEET)):
                                part_df_export = df_export.iloc[start_row_part : start_row_part + MAX_ROWS_PER_SHEET]
                                try:
                                    part_df_export.to_excel(writer_export, sheet_name=f"{sheet_name_export[:25]}_Part{i_part+1}", index=False) 
                                except Exception as e_ex_part: st.error(f"Error writing part {sheet_name_export}_Part{i_part+1}: {e_ex_part}")
                        else:
                            try:
                                df_export.to_excel(writer_export, sheet_name=sheet_name_export, index=False)
                            except Exception as e_ex_sheet: st.error(f"Error writing sheet {sheet_name_export}: {e_ex_sheet}")
                
                if export_data_buffer.tell() > 0: 
                    export_data_buffer.seek(0)
                    wb_loaded_export = load_workbook(export_data_buffer)
                    for sn_in_file in wb_loaded_export.sheetnames:
                        ws_export = wb_loaded_export[sn_in_file]
                        base_match = re.match(r"^(.*?)_Part\d+$", sn_in_file)
                        orig_df_key = base_match.group(1) if base_match else sn_in_file
                        
                        if orig_df_key not in workbook or workbook[orig_df_key].empty: continue 
                        
                        df_for_widths = workbook[orig_df_key] 

                        for col_idx_excel, column_cells in enumerate(ws_export.columns, 1):
                            if col_idx_excel -1 >= len(df_for_widths.columns): 
                                continue
                            
                            col_name_df = df_for_widths.columns[col_idx_excel-1]
                            max_l = len(str(col_name_df)) 

                            if col_name_df in df_for_widths and not df_for_widths[col_name_df].empty:
                                lengths = df_for_widths[col_name_df].astype(str).str.len()
                                if lengths.notna().any():
                                    max_l = max(max_l, int(lengths.max(skipna=True)))
                            
                            try:
                                ws_export.column_dimensions[get_column_letter(col_idx_excel)].width = max_l + 2
                            except Exception as e_width_final:
                                st.warning(f"Error setting width for column '{col_name_df}' in sheet '{sn_in_file}': {e_width_final}")
                    
                    final_export_bytes = io.BytesIO()
                    wb_loaded_export.save(final_export_bytes)
                    st.session_state.excel_data_to_download = final_export_bytes.getvalue()
                    st.session_state.data_processed = True
                    st.session_state.show_success_message = True
                else: 
                    if not any(not df.empty for df in workbook.values()): 
                         st.info("Workbook is empty. Nothing to export.")
                         st.session_state.excel_data_to_download = None 
                         st.session_state.data_processed = True 
                         st.session_state.show_success_message = False 
                    else: 
                        st.error("Failed to create a valid export file, buffer is empty despite processed data.")
                        st.session_state.error_occurred = True


        except Exception as e_main_process:
            st.session_state.error_occurred = True
            status_text.text("") 
            st.error(f"❌ An critical error occurred during processing: {e_main_process}")
            st.exception(e_main_process) 
            st.session_state.excel_data_to_download = None
            st.session_state.data_processed = False 
        
        finally:
            prog_bar_global.empty() 
            if not st.session_state.get('show_success_message', False) and not st.session_state.error_occurred :
                status_text_global.text("")


if st.session_state.get('show_success_message', False) and st.session_state.get('excel_data_to_download') is not None and not st.session_state.get('error_occurred', False):
    status_text_global.empty() 
    prog_bar_global.empty()    
    st.success("✅ Workbook processing complete! Click below to download.")
    st.download_button(
        "⬇️ Download Corrected Workbook",
        data=st.session_state.excel_data_to_download,
        file_name=st.session_state.out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key='final_download_button_main'
    )
    st.session_state.show_success_message = False 

elif st.session_state.get('error_occurred', False) and file is not None: 
    prog_bar_global.empty() 
    status_text_global.text("Processing failed. Please check errors reported above.")
